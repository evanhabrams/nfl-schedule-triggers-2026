"""
build_triggers.py — Generate triggers.json for the NFL Schedule Triggers tool.

Usage:
    python3 build_triggers.py --season 2025
    python3 build_triggers.py --season 2026   (once 2026 data is in the Triggers workbook)

Reads:
    - NFL_Schedule__NEW___12_.xlsx  → "Brand Logos" tab (team colors, abbreviations, logo URLs)
    - NFL_Schedule_Triggers__2026_.xlsx → "F1 Triggers {season}" tab (game-by-game trigger fires)

Writes:
    - triggers.json
"""
import argparse
from openpyxl import load_workbook
from collections import defaultdict
from datetime import datetime
import json
import sys

MASTER = "/mnt/user-data/uploads/NFL_Schedule__NEW___15_.xlsx"
TRIG   = "/mnt/user-data/uploads/NFL_Schedule_Triggers__2026___3_.xlsx"

# Current franchises only — skip historical rows (OAK, SD, LA, STL)
CURRENT_ABBRS = {'ARI','ATL','BAL','BUF','CAR','CHI','CIN','CLE','DAL','DEN',
                 'DET','GB','HOU','IND','JAX','KC','LAR','LAC','LV','MIA',
                 'MIN','NE','NO','NYG','NYJ','PHI','PIT','SEA','SF','TB','TEN','WAS'}
# Triggers-sheet name → Brand Logos canonical
NAME_MAP = {'Niners': '49ers'}

# Trigger definitions — Early Hot Host/Road dropped (1 fire/season, not a real signal).
# Framework is now 6 positives + 8 negatives. Negatives renumbered (no N7b gap).
TRIGGERS = [
    {'code':'P1', 'label':'Rest +3',       'polarity':'P', 'desc':'Focal Days Rest - Opp Days Rest >= 3'},
    {'code':'P2', 'label':'TZ +3',         'polarity':'P', 'desc':'|Opp Travel \u0394| - |Focal Travel \u0394| >= 3'},
    {'code':'P3', 'label':'Open \u2192 Dome',   'polarity':'P', 'desc':'Focal prev game roof = Open AND this = Fixed/Retractable'},
    {'code':'P4', 'label':'3+ Hm',         'polarity':'P', 'desc':"Focal's 3rd+ consecutive Home game (bye breaks streak)"},
    {'code':'P5', 'label':'Cold v Hot',    'polarity':'P', 'desc':'Focal home in Cold venue; opp Warm/Dome; Date >= Nov 15'},
    {'code':'P6', 'label':'Opp Off Intl',  'polarity':'P', 'desc':"Opp's prev game in GMT/CET/BRT (international)"},
    {'code':'N1', 'label':'Rest -3',       'polarity':'N', 'desc':'Opp Days Rest - Focal Days Rest >= 3'},
    {'code':'N2', 'label':'TZ -3',         'polarity':'N', 'desc':'|Focal Travel \u0394| - |Opp Travel \u0394| >= 3'},
    {'code':'N3', 'label':'Dome \u2192 Open',   'polarity':'N', 'desc':'Focal prev = Fixed/Retractable AND this = Open'},
    {'code':'N4', 'label':'3+ Rd',         'polarity':'N', 'desc':"Focal's 3rd+ consecutive Away+Neutral (bye breaks streak)"},
    {'code':'N5', 'label':'Off Intl',      'polarity':'N', 'desc':"Focal's prev game in GMT/CET/BRT"},
    {'code':'N6', 'label':'Off B2B Div',   'polarity':'N', 'desc':'Previous 2 games both vs division opponents (no bye)'},
    {'code':'N7', 'label':'Hot v Cold',    'polarity':'N', 'desc':'Focal away (Warm/Dome) at Cold venue; Date >= Nov 15'},
    {'code':'N8', 'label':'@ DEN',         'polarity':'N', 'desc':'Focal away at Denver'},
]

# Map workbook trigger-column codes → current codes (workbook still has old P5a/N7a/N7c columns).
# Old Early Hot columns (P5b, N7b) are intentionally absent → those fires are ignored.
WORKBOOK_CODE_MAP = {
    'P1':'P1', 'P2':'P2', 'P3':'P3', 'P4':'P4', 'P5a':'P5', 'P6':'P6',
    'N1':'N1', 'N2':'N2', 'N3':'N3', 'N4':'N4', 'N5':'N5', 'N6':'N6',
    'N7a':'N7', 'N7c':'N8',
}


def build(season, out_path):
    # ---- Team meta from Brand Logos ----
    wb_m = load_workbook(MASTER, data_only=True)
    ws_bl = wb_m["Brand Logos"]
    H = {ws_bl.cell(row=1, column=c).value: c for c in range(1, ws_bl.max_column+1)}

    teams = {}
    for r in range(2, ws_bl.max_row + 1):
        abbr = ws_bl.cell(row=r, column=H['team_abbr']).value
        if abbr not in CURRENT_ABBRS:
            continue
        nick = ws_bl.cell(row=r, column=H['team_nick']).value
        teams[nick] = {
            'abbr': abbr,
            'name': ws_bl.cell(row=r, column=H['team_name']).value,
            'color': ws_bl.cell(row=r, column=H['team_color']).value,
            'color2': ws_bl.cell(row=r, column=H['team_color2']).value,
            'conf': ws_bl.cell(row=r, column=H['team_conf']).value,
            'div': ws_bl.cell(row=r, column=H['team_division']).value,
            # ESPN logo (production — works on GitHub Pages / real browsers)
            'logo': ws_bl.cell(row=r, column=H['team_logo_espn']).value,
            # jsDelivr-proxied nflverse squared logo (works in more preview surfaces)
            'logoAlt': f"https://cdn.jsdelivr.net/gh/nflverse/nflverse-pbp@master/squared_logos/{abbr}.png",
        }
    if len(teams) != 32:
        print(f"WARNING: expected 32 teams, got {len(teams)}", file=sys.stderr)

    # ---- Game-level trigger fires ----
    sheet_name = f"F1 Triggers {season}"
    wb_t = load_workbook(TRIG, data_only=True)
    if sheet_name not in wb_t.sheetnames:
        print(f"ERROR: '{sheet_name}' not found in Triggers workbook. "
              f"Available: {wb_t.sheetnames}", file=sys.stderr)
        sys.exit(1)
    ws_t = wb_t[sheet_name]
    th = {str(ws_t.cell(row=1, column=c).value).replace('\n','|'): c
          for c in range(1, ws_t.max_column+1) if ws_t.cell(row=1, column=c).value}
    trig_col = {}
    for k, v in th.items():
        if '|' in k and not k.startswith('F1|'):
            trig_col[k.split('|', 1)[0]] = v

    fires = defaultdict(lambda: defaultdict(list))
    # track distinct weeks per polarity per team for "games with triggers"
    pos_weeks = defaultdict(set)
    neg_weeks = defaultdict(set)
    tot = defaultdict(lambda: {'pos': 0, 'neg': 0})

    for r in range(2, ws_t.max_row + 1):
        team_raw = ws_t.cell(row=r, column=th['Team']).value
        if not team_raw:
            continue
        team = NAME_MAP.get(team_raw, team_raw)
        opp = NAME_MAP.get(ws_t.cell(row=r, column=th['Opp']).value,
                           ws_t.cell(row=r, column=th['Opp']).value)
        week = int(ws_t.cell(row=r, column=th['Week']).value)
        han = ws_t.cell(row=r, column=th['H/A/N']).value
        date_val = ws_t.cell(row=r, column=th['Date']).value
        date_iso = date_val.strftime('%Y-%m-%d') if isinstance(date_val, datetime) else None

        # Read raw workbook codes, translate to current codes via WORKBOOK_CODE_MAP.
        # Workbook codes not in the map (P5b, N7b = Early Hot) are dropped here.
        fired = []
        for wb_code, col in trig_col.items():
            if ws_t.cell(row=r, column=col).value != 1:
                continue
            mapped = WORKBOOK_CODE_MAP.get(wb_code)
            if mapped:
                fired.append(mapped)
        if not fired:
            continue

        # Net is recomputed from the mapped (post-drop) triggers, not the workbook's F1 columns,
        # so it reflects the 6-pos / 8-neg framework.
        net = sum(1 for c in fired if c.startswith('P')) - sum(1 for c in fired if c.startswith('N'))
        game = {'week': week, 'date': date_iso, 'opp': opp, 'han': han,
                'triggers': fired, 'net': net}

        for code in fired:
            fires[team][code].append(game)
            if code.startswith('P'):
                tot[team]['pos'] += 1
                pos_weeks[team].add(week)
            else:
                tot[team]['neg'] += 1
                neg_weeks[team].add(week)

    # ---- Assemble ----
    out = {'season': season, 'teams': teams, 'triggers': TRIGGERS, 'fires': {}}
    for team in teams:
        out['fires'][team] = {
            'totals': {
                'pos': tot[team]['pos'],
                'neg': tot[team]['neg'],
                'net': tot[team]['pos'] - tot[team]['neg'],
                'posGames': len(pos_weeks[team]),   # distinct games with >=1 positive trigger
                'negGames': len(neg_weeks[team]),   # distinct games with >=1 negative trigger
            },
            'byTrigger': {code: {'count': len(g), 'games': g}
                          for code, g in fires[team].items()},
        }

    with open(out_path, 'w') as f:
        json.dump(out, f, indent=2, default=str)
    print(f"Wrote {out_path} — season {season}, {len(teams)} teams")


if __name__ == '__main__':
    ap = argparse.ArgumentParser()
    ap.add_argument('--season', type=int, default=2025)
    ap.add_argument('--out', default='/home/claude/triggers_tool/triggers.json')
    args = ap.parse_args()
    build(args.season, args.out)
